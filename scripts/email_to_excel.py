#!/usr/bin/env python3
"""
Email Topic Monitor - Excel Output Generator

Takes a JSON file of categorized emails and produces a formatted Excel workbook
with one sheet per topic, urgency highlighting, and a summary sheet.

Usage:
    python3 email_to_excel.py <input_json> [--output-dir PATH]

Input JSON format:
[
    {
        "topic": "Customer Complaints",
        "sender": "jane@acme.com",
        "subject": "Issue with order #123",
        "date": "2026-04-14T19:03:58Z",
        "snippet": "I'm having a problem with...",
        "thread_id": "19d8d61075f498a2",
        "urgent": true,
        "urgency_reason": "Customer threatening to cancel"
    },
    ...
]
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.views import BookView


def parse_args():
    parser = argparse.ArgumentParser(description="Generate Excel from categorized emails")
    parser.add_argument("input_json", help="Path to JSON file with categorized emails")
    parser.add_argument("--output-dir", help="Output directory", default=".")
    parser.add_argument("--slug", help="Optional slug for output filename", default="")
    return parser.parse_args()


# -- Template-derived style constants --
FONT_NAME = "Calisto MT"
FONT_SIZE = 12
FONT_SIZE_TITLE = 14
FONT_SIZE_HEADER = 12
MIN_ROW_HEIGHT = 26
ZOOM_SCALE = 130
WINDOW_WIDTH = 25420
WINDOW_HEIGHT = 14040
WINDOW_X = 460
WINDOW_Y = 1380

# Columns where values are short and should be centered horizontally
CENTER_COLUMNS = {"Date", "Topic", "Urgent", "Urgency Reason", "Count", "Latest Date", "Thread ID"}

# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=FONT_SIZE_HEADER)
BODY_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=FONT_SIZE_TITLE)
SUBTITLE_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
TOPIC_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=FONT_SIZE)
URGENT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
URGENT_FONT = Font(name=FONT_NAME, color="9C0006", bold=True, size=FONT_SIZE)
TOPIC_FILLS = {
    0: PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),  # blue
    1: PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # green
    2: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # orange
    3: PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),  # gray
    4: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # yellow
}
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def cell_alignment(col_name, wrap=False):
    """Return alignment based on column type."""
    h_align = "center" if col_name in CENTER_COLUMNS else None
    return Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)


def apply_sheet_view(ws):
    """Set zoom and view properties on a sheet."""
    ws.sheet_view.zoomScale = ZOOM_SCALE


def set_row_height(ws, row_idx):
    """Ensure minimum row height."""
    ws.row_dimensions[row_idx].height = MIN_ROW_HEIGHT


def write_header(ws, columns, row=1):
    """Write styled header row."""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    set_row_height(ws, row)
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def write_email_row(ws, row_idx, email, columns):
    """Write one email row with optional urgency highlighting."""
    values = {
        "Date": email.get("date", ""),
        "Sender": email.get("sender", ""),
        "Subject": email.get("subject", ""),
        "Snippet": email.get("snippet", ""),
        "Urgent": "YES" if email.get("urgent") else "",
        "Urgency Reason": email.get("urgency_reason", "") if email.get("urgent") else "",
        "Topic": email.get("topic", "Uncategorized"),
        "Thread ID": email.get("thread_id", ""),
    }

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=values.get(col_name, ""))
        cell.border = THIN_BORDER
        cell.font = BODY_FONT
        cell.alignment = cell_alignment(col_name, wrap=(col_name in ("Snippet", "Subject")))

        if email.get("urgent"):
            cell.fill = URGENT_FILL
            if col_name in ("Urgent", "Subject"):
                cell.font = URGENT_FONT

    set_row_height(ws, row_idx)


def set_column_widths(ws, columns):
    """Set reasonable column widths."""
    widths = {
        "Date": 22,
        "Sender": 32,
        "Subject": 50,
        "Snippet": 60,
        "Urgent": 9,
        "Urgency Reason": 35,
        "Topic": 26,
        "Thread ID": 12,
    }
    for col_idx, col_name in enumerate(columns, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(col_name, 20)


def create_workbook(emails, output_path):
    """Create the full Excel workbook with summary + per-topic sheets."""
    wb = Workbook()

    # Set workbook window size (not full screen)
    wb.views = [BookView(
        xWindow=WINDOW_X,
        yWindow=WINDOW_Y,
        windowWidth=WINDOW_WIDTH,
        windowHeight=WINDOW_HEIGHT,
    )]

    # Group emails by topic
    by_topic = {}
    for email in emails:
        topic = email.get("topic", "Uncategorized")
        by_topic.setdefault(topic, []).append(email)

    # Sort each topic's emails chronologically
    for topic in by_topic:
        by_topic[topic].sort(key=lambda e: e.get("date", ""))

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    apply_sheet_view(ws_summary)

    cell = ws_summary.cell(row=1, column=1, value="Email Topic Monitor Report")
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    set_row_height(ws_summary, 1)

    cell = ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(vertical="center")
    set_row_height(ws_summary, 2)

    cell = ws_summary.cell(row=3, column=1, value=f"Total emails: {len(emails)}")
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(vertical="center")
    set_row_height(ws_summary, 3)

    urgent_count = sum(1 for e in emails if e.get("urgent"))
    if urgent_count:
        cell = ws_summary.cell(row=4, column=1, value=f"URGENT items: {urgent_count}")
        cell.font = URGENT_FONT
        cell.fill = URGENT_FILL
        cell.alignment = Alignment(vertical="center")
    set_row_height(ws_summary, 4)

    # Topic summary table
    summary_headers = ["Topic", "Count", "Urgent", "Latest Date"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=6, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    set_row_height(ws_summary, 6)

    for row_idx, (topic, topic_emails) in enumerate(sorted(by_topic.items()), 7):
        topic_urgent = sum(1 for e in topic_emails if e.get("urgent"))
        latest = max(e.get("date", "") for e in topic_emails)
        values = [topic, len(topic_emails), topic_urgent if topic_urgent else "", latest]
        for col_idx, val in enumerate(values, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            col_name = summary_headers[col_idx - 1]
            cell.alignment = cell_alignment(col_name)
        set_row_height(ws_summary, row_idx)

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.column_dimensions["D"].width = 22

    # --- All Emails sheet (chronological) ---
    all_columns = ["Date", "Topic", "Sender", "Subject", "Snippet", "Urgent", "Urgency Reason"]
    ws_all = wb.create_sheet("All Emails")
    apply_sheet_view(ws_all)
    write_header(ws_all, all_columns)
    all_sorted = sorted(emails, key=lambda e: e.get("date", ""))
    for row_idx, email in enumerate(all_sorted, 2):
        write_email_row(ws_all, row_idx, email, all_columns)
    set_column_widths(ws_all, all_columns)
    ws_all.auto_filter.ref = ws_all.dimensions

    # --- Per-topic sheets ---
    topic_columns = ["Date", "Sender", "Subject", "Snippet", "Urgent", "Urgency Reason"]
    for topic_idx, (topic, topic_emails) in enumerate(sorted(by_topic.items())):
        # Sheet name max 31 chars, no special chars
        sheet_name = re.sub(r'[^\w\s-]', '', topic)[:31]
        ws_topic = wb.create_sheet(sheet_name)
        apply_sheet_view(ws_topic)

        # Topic header with color
        fill = TOPIC_FILLS.get(topic_idx % len(TOPIC_FILLS))
        cell = ws_topic.cell(row=1, column=1, value=f"{topic} ({len(topic_emails)} emails)")
        cell.font = TOPIC_TITLE_FONT
        cell.alignment = Alignment(vertical="center")
        set_row_height(ws_topic, 1)
        set_row_height(ws_topic, 2)

        # Column headers at row 3
        write_header(ws_topic, topic_columns, row=3)

        for row_idx, email in enumerate(topic_emails, 4):
            write_email_row(ws_topic, row_idx, email, topic_columns)

        set_column_widths(ws_topic, topic_columns)

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"  {len(emails)} emails across {len(by_topic)} topics")
    print(f"  {urgent_count} urgent items highlighted")
    print(f"  Sheets: Summary, All Emails, + {len(by_topic)} topic sheets")


def main():
    args = parse_args()

    with open(args.input_json) as f:
        emails = json.load(f)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = f"_{args.slug}" if args.slug else ""
    filename = f"email_monitor_{timestamp}{slug}.xlsx"
    output_path = os.path.join(args.output_dir, filename)

    os.makedirs(args.output_dir, exist_ok=True)
    create_workbook(emails, output_path)


if __name__ == "__main__":
    main()
