#!/usr/bin/env python3
"""
Transaction Data Analyzer - Multi-step pipeline with checkpointed outputs.

Each step saves its output before proceeding, so partial runs are still useful.

Usage:
    python3 analyze_data.py <file_path> [options]

Options:
    --start YYYY-MM-DD    Start date for filtering
    --end YYYY-MM-DD      End date for filtering
    --days N              Number of days from start date
    --before YYYY-MM-DD   Show all transactions before this date
    --after YYYY-MM-DD    Show all transactions after this date
    --output-dir PATH     Override output directory
"""

import argparse
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib
matplotlib.use("Agg")  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(description="Analyze transaction data")
    parser.add_argument("file_path", help="Path to CSV or Excel file")
    parser.add_argument("--start", help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end", help="End date (YYYY-MM-DD)")
    parser.add_argument("--days", type=int, help="Number of days from start date")
    parser.add_argument("--before", help="Show transactions before date (YYYY-MM-DD)")
    parser.add_argument("--after", help="Show transactions after date (YYYY-MM-DD)")
    parser.add_argument("--output-dir", help="Override output directory")
    return parser.parse_args()


def resolve_date_range(args):
    """Return (start_date, end_date, has_prior_period) from args."""
    if args.start and args.end:
        return pd.Timestamp(args.start), pd.Timestamp(args.end), True
    if args.start and args.days:
        start = pd.Timestamp(args.start)
        return start, start + timedelta(days=args.days - 1), True
    if args.before:
        return None, pd.Timestamp(args.before), False
    if args.after:
        return pd.Timestamp(args.after), None, False
    return None, None, False


# ---------------------------------------------------------------------------
# Step 1: Ingest & Profile
# ---------------------------------------------------------------------------

def load_data(file_path: str) -> pd.DataFrame:
    """Load CSV or Excel file into a DataFrame."""
    p = Path(file_path)
    if not p.exists():
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)

    ext = p.suffix.lower()
    if ext == ".csv":
        df = pd.read_csv(file_path, on_bad_lines="skip")
    elif ext in (".xlsx", ".xls"):
        sheets = pd.read_excel(file_path, sheet_name=None)
        if len(sheets) == 1:
            df = next(iter(sheets.values()))
        else:
            df = pd.concat(sheets.values(), ignore_index=True)
            print(f"Concatenated {len(sheets)} sheets: {list(sheets.keys())}")
    else:
        print(f"ERROR: Unsupported file format: {ext}")
        sys.exit(1)

    print(f"Loaded {len(df)} rows, {len(df.columns)} columns from {p.name}")
    return df


def detect_date_column(df: pd.DataFrame) -> str | None:
    """Find the most likely date column."""
    for col in df.columns:
        if df[col].dtype == "datetime64[ns]":
            return col
        col_lower = col.lower()
        if any(kw in col_lower for kw in ["date", "time", "day", "period", "wkend"]):
            try:
                parsed = pd.to_datetime(df[col], errors="coerce")
                if parsed.notna().sum() > len(df) * 0.8:
                    return col
            except Exception:
                continue
    return None


def detect_amount_columns(df: pd.DataFrame) -> list[str]:
    """Find numeric columns likely representing monetary amounts."""
    amount_keywords = ["amount", "price", "total", "revenue", "cost", "spend",
                       "charge", "discount", "subtot", "net", "gross"]
    result = []
    for col in df.columns:
        col_lower = col.lower()
        if any(kw in col_lower for kw in amount_keywords):
            result.append(col)
        elif df[col].dtype in ("float64", "int64") and col not in result:
            # Check if values look monetary (not IDs, not counts)
            vals = df[col].dropna()
            if len(vals) > 0:
                if vals.abs().median() > 0.5 and vals.abs().median() < 1_000_000:
                    if col.lower() not in ("seqn", "id", "quantity", "impressions"):
                        result.append(col)
    return result


def detect_category_columns(df: pd.DataFrame) -> list[str]:
    """Find categorical columns useful for grouping.

    Cap at 500 uniques: downstream summaries take the top-N. The lower 50-uniques
    cap was too aggressive — a column literally named ``Category`` with 248
    distinct values is a genuine grouping column, not noise.
    """
    cat_keywords = ["category", "type", "channel", "status", "account", "group",
                    "class", "department", "condition"]
    result = []
    for col in df.columns:
        col_lower = col.lower().replace("_", " ")
        if any(kw in col_lower for kw in cat_keywords):
            nunique = df[col].nunique()
            if 2 <= nunique <= 500:
                result.append(col)
    return result


def profile_data(df: pd.DataFrame, date_col: str | None,
                 amount_cols: list[str], category_cols: list[str]) -> str:
    """Generate a text profile of the dataset."""
    lines = ["# Data Profile\n"]
    lines.append(f"- **Rows:** {len(df):,}")
    lines.append(f"- **Columns:** {len(df.columns)}")
    lines.append(f"- **Date column:** {date_col or 'Not detected'}")
    lines.append(f"- **Amount/value columns:** {', '.join(amount_cols) or 'None detected'}")
    lines.append(f"- **Category columns:** {', '.join(category_cols) or 'None detected'}")

    if date_col:
        dates = pd.to_datetime(df[date_col], errors="coerce").dropna()
        lines.append(f"- **Date range:** {dates.min().strftime('%Y-%m-%d')} to {dates.max().strftime('%Y-%m-%d')}")

    lines.append("\n## Column Summary\n")
    lines.append("| Column | Type | Unique | Nulls | Sample Values |")
    lines.append("|--------|------|--------|-------|---------------|")
    for col in df.columns:
        dtype = str(df[col].dtype)
        nunique = df[col].nunique()
        nulls = df[col].isnull().sum()
        samples = df[col].dropna().unique()[:3]
        sample_str = ", ".join(str(s)[:30] for s in samples)
        lines.append(f"| {col} | {dtype} | {nunique:,} | {nulls:,} | {sample_str} |")

    # Amount statistics
    if amount_cols:
        lines.append("\n## Amount Statistics\n")
        for col in amount_cols:
            vals = pd.to_numeric(df[col].astype(str).str.replace(",", ""), errors="coerce")
            if vals.notna().sum() > 0:
                lines.append(f"### {col}")
                lines.append(f"- Count: {vals.notna().sum():,}")
                lines.append(f"- Sum: {vals.sum():,.2f}")
                lines.append(f"- Mean: {vals.mean():,.2f}")
                lines.append(f"- Median: {vals.median():,.2f}")
                lines.append(f"- Min: {vals.min():,.2f} / Max: {vals.max():,.2f}")
                lines.append(f"- Std Dev: {vals.std():,.2f}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Step 2: Filter & Select Key Fields
# ---------------------------------------------------------------------------

def clean_amount_column(df: pd.DataFrame, col: str) -> pd.Series:
    """Convert an amount column to numeric, handling commas and strings."""
    return pd.to_numeric(df[col].astype(str).str.replace(",", ""), errors="coerce")


def filter_by_date(df: pd.DataFrame, date_col: str,
                   start: pd.Timestamp | None, end: pd.Timestamp | None) -> pd.DataFrame:
    """Filter DataFrame by date range."""
    dates = pd.to_datetime(df[date_col], errors="coerce")
    mask = dates.notna()
    if start is not None:
        mask = mask & (dates >= start)
    if end is not None:
        mask = mask & (dates <= end)
    return df[mask].copy()


def select_key_fields(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Drop columns unlikely to be useful for analysis.

    Returns (filtered_df, dropped_column_names).
    """
    drop_cols = []
    for col in df.columns:
        # Drop if all values are the same
        if df[col].nunique() <= 1:
            drop_cols.append(col)
            continue
        # Drop if >80% null
        if df[col].isnull().sum() > len(df) * 0.8:
            drop_cols.append(col)
            continue
        # Drop address-like columns, long text IDs
        col_lower = col.lower()
        if any(kw in col_lower for kw in ["address", "bill_addr", "ship_addr"]):
            drop_cols.append(col)
            continue
        # Drop columns where average string length > 80 (likely descriptions/addresses)
        if df[col].dtype == "object":
            avg_len = df[col].dropna().astype(str).str.len().mean()
            if avg_len > 80:
                drop_cols.append(col)
                continue

    kept = df.drop(columns=drop_cols)
    return kept, drop_cols


def write_filtered_excel(df: pd.DataFrame, output_path: str, date_col: str | None):
    """Write filtered data to a nicely formatted Excel file."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Data"

    # Write headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Format numbers
            if isinstance(value, float):
                cell.number_format = "#,##0.00"

    # Auto-width columns (approximate)
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), df.iloc[:, col_idx - 1].astype(str).str.len().median())
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add autofilter
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"  Saved filtered Excel: {output_path} ({len(df)} rows, {len(df.columns)} columns)")


# ---------------------------------------------------------------------------
# Step 3: Analyze & Chart
# ---------------------------------------------------------------------------

def compute_summary(df: pd.DataFrame, amount_cols: list[str],
                    category_cols: list[str], date_col: str | None,
                    prior_df: pd.DataFrame | None = None) -> str:
    """Generate a markdown summary report."""
    lines = ["# Analysis Summary\n"]

    lines.append(f"**Records in period:** {len(df):,}")
    if prior_df is not None:
        lines.append(f"**Records in prior period:** {len(prior_df):,}")
    lines.append("")

    # Amount summaries
    for col in amount_cols:
        vals = clean_amount_column(df, col)
        if vals.notna().sum() == 0:
            continue
        lines.append(f"## {col}\n")
        lines.append(f"| Metric | Current Period |" + (" Prior Period | Change |" if prior_df is not None else ""))
        lines.append(f"|--------|---------------|" + ("--------------|--------|" if prior_df is not None else ""))

        metrics = {
            "Total": vals.sum(),
            "Mean": vals.mean(),
            "Median": vals.median(),
            "Min": vals.min(),
            "Max": vals.max(),
            "Count": vals.notna().sum(),
        }

        if prior_df is not None:
            prior_vals = clean_amount_column(prior_df, col)
            prior_metrics = {
                "Total": prior_vals.sum(),
                "Mean": prior_vals.mean(),
                "Median": prior_vals.median(),
                "Min": prior_vals.min(),
                "Max": prior_vals.max(),
                "Count": prior_vals.notna().sum(),
            }
        else:
            prior_metrics = {}

        for metric, val in metrics.items():
            fmt = f"{val:,.2f}" if metric != "Count" else f"{val:,.0f}"
            row = f"| {metric} | {fmt} |"
            if prior_df is not None and metric in prior_metrics:
                pval = prior_metrics[metric]
                pfmt = f"{pval:,.2f}" if metric != "Count" else f"{pval:,.0f}"
                if pval != 0:
                    change = ((val - pval) / abs(pval)) * 100
                    row += f" {pfmt} | {change:+.1f}% |"
                else:
                    row += f" {pfmt} | N/A |"
            lines.append(row)
        lines.append("")

    # Category breakdowns
    for col in category_cols:
        if col not in df.columns:
            continue
        lines.append(f"## Distribution by {col}\n")
        counts = df[col].value_counts().head(15)
        lines.append("| Category | Count | % |")
        lines.append("|----------|-------|---|")
        for cat, count in counts.items():
            pct = count / len(df) * 100
            lines.append(f"| {cat} | {count:,} | {pct:.1f}% |")
        lines.append("")

    return "\n".join(lines)


def generate_charts(df: pd.DataFrame, amount_cols: list[str],
                    category_cols: list[str], date_col: str | None,
                    output_dir: str, prior_df: pd.DataFrame | None = None) -> list[str]:
    """Generate charts and return list of saved file paths."""
    chart_files = []
    plt.style.use("seaborn-v0_8-whitegrid")

    primary_amount = amount_cols[0] if amount_cols else None

    # Chart 1: Category breakdown (bar chart)
    if primary_amount and category_cols:
        cat_col = category_cols[0]
        vals = clean_amount_column(df, primary_amount)
        temp = df.copy()
        temp["_amount"] = vals
        grouped = temp.groupby(cat_col)["_amount"].sum().sort_values()

        # Show top 12 categories
        if len(grouped) > 12:
            top = grouped.tail(12)
        else:
            top = grouped

        fig, ax = plt.subplots(figsize=(10, 6))
        colors = plt.cm.RdYlGn(np.linspace(0.2, 0.8, len(top)))
        bars = ax.barh(range(len(top)), top.values, color=colors)
        ax.set_yticks(range(len(top)))
        ax.set_yticklabels(top.index, fontsize=9)
        ax.set_xlabel(primary_amount)
        ax.set_title(f"Total {primary_amount} by {cat_col}")
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f"${x:,.0f}" if abs(x) >= 1 else f"{x:,.2f}"))
        plt.tight_layout()
        path = os.path.join(output_dir, "step3_chart_by_category.png")
        fig.savefig(path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        chart_files.append(path)
        print(f"  Saved chart: {path}")

    # Chart 2: Trend over time
    if primary_amount and date_col and date_col in df.columns:
        vals = clean_amount_column(df, primary_amount)
        dates = pd.to_datetime(df[date_col], errors="coerce")
        temp = pd.DataFrame({"date": dates, "amount": vals}).dropna()

        if len(temp) > 0:
            # Determine appropriate grouping
            date_span = (temp["date"].max() - temp["date"].min()).days
            if date_span > 365:
                temp["period"] = temp["date"].dt.to_period("M")
                period_label = "Month"
            elif date_span > 60:
                temp["period"] = temp["date"].dt.to_period("W")
                period_label = "Week"
            else:
                temp["period"] = temp["date"].dt.to_period("D")
                period_label = "Day"

            trend = temp.groupby("period")["amount"].agg(["sum", "count"])

            fig, ax1 = plt.subplots(figsize=(12, 5))
            x = range(len(trend))
            ax1.bar(x, trend["sum"], alpha=0.7, color="#4472C4", label=f"Total {primary_amount}")
            ax1.set_ylabel(f"Total {primary_amount}", color="#4472C4")
            ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f"${x:,.0f}"))

            ax2 = ax1.twinx()
            ax2.plot(x, trend["count"], color="#ED7D31", linewidth=2, marker="o",
                     markersize=4, label="Transaction Count")
            ax2.set_ylabel("Count", color="#ED7D31")

            # X-axis labels
            labels = [str(p) for p in trend.index]
            step = max(1, len(labels) // 15)
            ax1.set_xticks(x[::step])
            ax1.set_xticklabels(labels[::step], rotation=45, ha="right", fontsize=8)
            ax1.set_title(f"{primary_amount} Trend by {period_label}")

            lines1, labels1 = ax1.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left")

            plt.tight_layout()
            path = os.path.join(output_dir, "step3_chart_trend.png")
            fig.savefig(path, dpi=150, bbox_inches="tight")
            plt.close(fig)
            chart_files.append(path)
            print(f"  Saved chart: {path}")

    # Chart 3: Prior period comparison (if applicable)
    if prior_df is not None and primary_amount and category_cols:
        cat_col = category_cols[0]
        curr_vals = clean_amount_column(df, primary_amount)
        prior_vals = clean_amount_column(prior_df, primary_amount)

        curr_temp = df.copy()
        curr_temp["_amount"] = curr_vals
        prior_temp = prior_df.copy()
        prior_temp["_amount"] = prior_vals

        curr_grouped = curr_temp.groupby(cat_col)["_amount"].sum()
        prior_grouped = prior_temp.groupby(cat_col)["_amount"].sum()

        # Top categories by current period
        all_cats = curr_grouped.abs().sort_values(ascending=False).head(10).index
        curr_top = curr_grouped.reindex(all_cats, fill_value=0)
        prior_top = prior_grouped.reindex(all_cats, fill_value=0)

        fig, ax = plt.subplots(figsize=(10, 6))
        x = np.arange(len(all_cats))
        width = 0.35
        ax.bar(x - width / 2, prior_top.values, width, label="Prior Period",
               color="#A5A5A5", alpha=0.8)
        ax.bar(x + width / 2, curr_top.values, width, label="Current Period",
               color="#4472C4", alpha=0.8)
        ax.set_xticks(x)
        ax.set_xticklabels(all_cats, rotation=45, ha="right", fontsize=8)
        ax.set_ylabel(primary_amount)
        ax.set_title(f"{primary_amount} by {cat_col}: Current vs Prior Period")
        ax.legend()
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f"${x:,.0f}"))
        plt.tight_layout()
        path = os.path.join(output_dir, "step3_chart_comparison.png")
        fig.savefig(path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        chart_files.append(path)
        print(f"  Saved chart: {path}")

    return chart_files


# ---------------------------------------------------------------------------
# Step 4: Outlier & Data Quality Detection
# ---------------------------------------------------------------------------

def detect_outliers_and_quality(df: pd.DataFrame, amount_cols: list[str],
                                date_col: str | None) -> str:
    """Detect outliers and data quality issues, return markdown report."""
    lines = ["# Data Quality Report\n"]
    issues_found = 0

    # Missing values
    lines.append("## Missing Values\n")
    missing = df.isnull().sum()
    has_missing = missing[missing > 0]
    if len(has_missing) > 0:
        lines.append("| Column | Missing | % |")
        lines.append("|--------|---------|---|")
        for col, count in has_missing.items():
            pct = count / len(df) * 100
            lines.append(f"| {col} | {count:,} | {pct:.1f}% |")
            issues_found += count
    else:
        lines.append("No missing values detected.\n")

    # Duplicate rows
    lines.append("\n## Duplicate Rows\n")
    dupes = df.duplicated().sum()
    if dupes > 0:
        lines.append(f"**{dupes:,} duplicate rows found** ({dupes / len(df) * 100:.1f}% of data)\n")
        issues_found += dupes
    else:
        lines.append("No duplicate rows detected.\n")

    # Amount outliers
    for col in amount_cols:
        vals = clean_amount_column(df, col)
        valid = vals.dropna()
        if len(valid) < 10:
            continue

        mean = valid.mean()
        std = valid.std()
        if std == 0:
            continue

        # Flag values beyond 3 standard deviations
        outlier_mask = (valid - mean).abs() > 3 * std
        outliers = valid[outlier_mask]

        lines.append(f"\n## Outliers in {col}\n")
        lines.append(f"Using 3-sigma threshold (mean={mean:,.2f}, std={std:,.2f})")
        lines.append(f"Threshold: < {mean - 3 * std:,.2f} or > {mean + 3 * std:,.2f}\n")

        if len(outliers) > 0:
            lines.append(f"**{len(outliers):,} outlier transactions found:**\n")
            issues_found += len(outliers)

            # Show up to 20 outliers
            outlier_rows = df.loc[outliers.index].head(20)
            display_cols = list(df.columns[:6])  # Show first 6 columns for context
            lines.append("| " + " | ".join(display_cols) + " |")
            lines.append("|" + "|".join(["---"] * len(display_cols)) + "|")
            for _, row in outlier_rows.iterrows():
                vals_str = [str(row.get(c, ""))[:30] for c in display_cols]
                lines.append("| " + " | ".join(vals_str) + " |")

            if len(outliers) > 20:
                lines.append(f"\n*... and {len(outliers) - 20} more outliers*")
        else:
            lines.append("No outliers detected.\n")

    # Date quality (if applicable)
    if date_col and date_col in df.columns:
        lines.append(f"\n## Date Quality ({date_col})\n")
        dates = pd.to_datetime(df[date_col], errors="coerce")
        bad_dates = dates.isna().sum() - df[date_col].isna().sum()  # Unparseable dates
        if bad_dates > 0:
            lines.append(f"**{bad_dates:,} unparseable date values found**\n")
            issues_found += bad_dates
        else:
            lines.append("All dates parsed successfully.\n")

    lines.append(f"\n---\n**Total issues flagged: {issues_found:,}**")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def main():
    args = parse_args()

    # Set up output directory
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.output_dir:
        output_dir = args.output_dir
    else:
        output_dir = os.path.join("output", f"analyze_data_{timestamp}")
    os.makedirs(output_dir, exist_ok=True)

    print(f"\n{'='*60}")
    print(f" Transaction Data Analyzer")
    print(f" Output: {output_dir}")
    print(f"{'='*60}\n")

    # Parse date range
    start_date, end_date, has_prior = resolve_date_range(args)
    if start_date or end_date:
        range_desc = []
        if start_date:
            range_desc.append(f"from {start_date.strftime('%Y-%m-%d')}")
        if end_date:
            range_desc.append(f"to {end_date.strftime('%Y-%m-%d')}")
        print(f"Date range: {' '.join(range_desc)}")
        if has_prior and start_date and end_date:
            period_days = (end_date - start_date).days
            prior_start = start_date - timedelta(days=period_days + 1)
            prior_end = start_date - timedelta(days=1)
            print(f"Prior period: {prior_start.strftime('%Y-%m-%d')} to {prior_end.strftime('%Y-%m-%d')}")
    print()

    # ===== STEP 1: Ingest & Profile =====
    print("STEP 1: Ingest & Profile")
    print("-" * 40)

    df = load_data(args.file_path)
    date_col = detect_date_column(df)
    amount_cols = detect_amount_columns(df)
    category_cols = detect_category_columns(df)

    # Ensure date column is datetime
    if date_col:
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

    # Clean amount columns
    for col in amount_cols:
        df[col] = clean_amount_column(df, col)

    profile = profile_data(df, date_col, amount_cols, category_cols)
    profile_path = os.path.join(output_dir, "step1_data_profile.md")
    with open(profile_path, "w") as f:
        f.write(profile)
    print(f"  Saved profile: {profile_path}")
    print(f"  Detected: date_col={date_col}, amounts={amount_cols}, categories={category_cols}")
    print()

    # ===== STEP 2: Filter & Select Key Fields =====
    print("STEP 2: Filter & Select Key Fields")
    print("-" * 40)

    if date_col and (start_date or end_date):
        filtered = filter_by_date(df, date_col, start_date, end_date)
        print(f"  Filtered: {len(df)} -> {len(filtered)} rows")
    else:
        filtered = df.copy()
        if not date_col:
            print("  No date column detected -- using all rows")
        else:
            print("  No date range specified -- using all rows")

    selected, dropped = select_key_fields(filtered)
    if dropped:
        print(f"  Dropped {len(dropped)} low-value columns: {dropped}")
    else:
        print("  All columns retained (none identified as low-value)")

    excel_path = os.path.join(output_dir, "step2_filtered_data.xlsx")
    write_filtered_excel(selected, excel_path, date_col)
    print()

    # ===== STEP 3: Analyze & Chart =====
    print("STEP 3: Analyze & Chart")
    print("-" * 40)

    # Get prior period data if applicable
    prior_df = None
    if has_prior and date_col and start_date and end_date:
        period_days = (end_date - start_date).days
        prior_start = start_date - timedelta(days=period_days + 1)
        prior_end = start_date - timedelta(days=1)
        prior_df = filter_by_date(df, date_col, prior_start, prior_end)
        print(f"  Prior period: {len(prior_df)} rows ({prior_start.strftime('%Y-%m-%d')} to {prior_end.strftime('%Y-%m-%d')})")

    # Filter amount and category cols to those still in selected
    active_amount_cols = [c for c in amount_cols if c in selected.columns]
    active_cat_cols = [c for c in category_cols if c in selected.columns]

    summary = compute_summary(selected, active_amount_cols, active_cat_cols,
                              date_col, prior_df)
    summary_path = os.path.join(output_dir, "step3_summary_report.md")
    with open(summary_path, "w") as f:
        f.write(summary)
    print(f"  Saved summary: {summary_path}")

    chart_files = generate_charts(selected, active_amount_cols, active_cat_cols,
                                  date_col, output_dir, prior_df)
    print(f"  Generated {len(chart_files)} charts")
    print()

    # ===== STEP 4: Outlier & Data Quality =====
    print("STEP 4: Outlier & Data Quality Detection")
    print("-" * 40)

    quality = detect_outliers_and_quality(selected, active_amount_cols, date_col)
    quality_path = os.path.join(output_dir, "step4_quality_report.md")
    with open(quality_path, "w") as f:
        f.write(quality)
    print(f"  Saved quality report: {quality_path}")
    print()

    # ===== Summary =====
    print(f"{'='*60}")
    print(f" COMPLETE -- All outputs in: {output_dir}")
    print(f"{'='*60}")
    print(f"  step1_data_profile.md     - Data profile and column summary")
    print(f"  step2_filtered_data.xlsx  - Filtered data ({len(selected)} rows, {len(selected.columns)} cols)")
    print(f"  step3_summary_report.md   - Analysis summary with statistics")
    for cf in chart_files:
        print(f"  {os.path.basename(cf):28s} - Chart")
    print(f"  step4_quality_report.md   - Outlier and quality flags")


if __name__ == "__main__":
    main()
